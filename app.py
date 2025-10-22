import io, base64, datetime as dt
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd

import plotly.graph_objects as go
from dash import Dash, dcc, html, Input, Output, State, callback_context
import dash_bootstrap_components as dbc

# Export PNG de Plotly
import plotly.io as pio
pio.kaleido.scope.default_format = "png"
pio.kaleido.scope.default_scale = 2

# Excel
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# =============== UTILITAIRES ===============
def lire_fichier_dBNext_from_content(content, filename):
    """Décodage upload + lecture CSV (col0=time, col1=LAeq)."""
    content_type, content_string = content.split(',')
    decoded = base64.b64decode(content_string)
    df = pd.read_csv(io.BytesIO(decoded), sep=None, engine="python", encoding="utf-8")
    if df.shape[1] < 2:
        raise ValueError(f"{filename}: besoin d’au moins 2 colonnes (temps + LAeq).")
    df = df.iloc[:, :2].copy()
    df.columns = ["time_block", "LAeq"]
    df["time_block"] = pd.to_datetime(df["time_block"], errors="coerce", dayfirst=True)
    df["LAeq"] = pd.to_numeric(df["LAeq"], errors="coerce")
    df = df.dropna(subset=["time_block", "LAeq"]).set_index("time_block").sort_index()
    return df

JOUR_IDX = {"lundi":0,"mardi":1,"mercredi":2,"jeudi":3,"vendredi":4,"samedi":5,"dimanche":6}

def normalize_jours(jours): return [JOUR_IDX[j.lower()] for j in (jours or []) if j.lower() in JOUR_IDX]

def build_intervals(t0, t1, h_deb, h_fin, jours=None):
    days = normalize_jours(jours)
    cur = dt.datetime(t0.year, t0.month, t0.day)
    out=[]
    while cur < t1:
        if (not days) or (cur.weekday() in days):
            if h_deb <= h_fin:
                s = cur.replace(hour=h_deb, minute=0, second=0)
                e = cur.replace(hour=h_fin, minute=0, second=0)
            else:
                s = cur.replace(hour=h_deb, minute=0, second=0)
                e = (cur + dt.timedelta(days=1)).replace(hour=h_fin, minute=0, second=0)
            if s < t1 and e > t0: out.append((max(s,t0), min(e,t1)))
        cur += dt.timedelta(days=1)
    return out

def concat_series(df, intervals, col="LAeq"):
    segs = [df.loc[(df.index>=s)&(df.index<=e), col] for s,e in intervals]
    return pd.concat(segs).sort_index() if segs else pd.Series(dtype=float)

def LAeq(series):
    if len(series)==0: return np.nan
    return float(10*np.log10(np.nanmean(10**(series/10))))

def LN(series, N):
    if len(series)==0: return np.nan
    return float(np.nanpercentile(series, 100-N))

def midnight_lines(tmin, tmax):
    start = pd.Timestamp(tmin).floor("D")
    end   = pd.Timestamp(tmax).ceil("D")
    return list(pd.date_range(start, end, freq="D"))

def hour_step(span_seconds):
    # pas horaire adaptatif pour l’axe du bas
    if span_seconds <= 6*3600:   return 3600000   # 1 h (ms)
    if span_seconds <= 24*3600:  return 2*3600000 # 2 h
    if span_seconds <= 3*86400:  return 3*3600000 # 3 h
    return 6*3600000                         # 6 h

def build_figure(df, removed_ts=None, diurne=(7,22), nocturne=(22,7)):
    """Trace Diurne/Nocturne, heures en bas, jours en haut, minuit, petites étiquettes."""
    if df.empty:
        fig = go.Figure()
        fig.add_annotation(text="Aucune donnée", showarrow=False, font=dict(size=16))
        return fig

    removed_ts = set(removed_ts or [])  # timestamps ISO à exclure
    # split séries
    t0, t1 = df.index.min(), df.index.max()
    all_intervals = {
        "Diurne":   build_intervals(t0, t1, diurne[0], diurne[1], ["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]),
        "Nocturne": build_intervals(t0, t1, nocturne[0], nocturne[1], ["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]),
    }
    series = {k: concat_series(df, v) for k,v in all_intervals.items()}

    colors = {"Diurne":"#1f77b4", "Nocturne":"#ff7f0e"}

    fig = go.Figure()

    # Traces (avec exclusion)
    for name, s in series.items():
        if s.empty: continue
        mask = ~s.index.astype(str).isin(removed_ts)
        s_clean = s[mask]
        fig.add_trace(go.Scatter(
            x=s_clean.index, y=s_clean.values,
            mode="lines", name=name, line=dict(width=1.2, color=colors.get(name, "#333")),
            hovertemplate="%{x|%d/%m/%Y %H:%M:%S}<br>LAeq=%{y:.1f} dB(A)<extra>"+name+"</extra>"
        ))
        # Points retirés (Perturbations)
        pert = s[~mask]
        if len(pert):
            fig.add_trace(go.Scatter(
                x=pert.index, y=pert.values, mode="markers", name=f"Perturbations {name}",
                marker=dict(size=6, color="red"), hovertemplate="%{x|%d/%m/%Y %H:%M:%S}<br>%{y:.1f} dB(A)<extra>Exclu</extra>"
            ))

    xmin, xmax = df.index.min(), df.index.max()
    span_s = (xmax - xmin).total_seconds()
    dtick_ms = hour_step(span_s)

    # Lignes de minuit
    shapes=[]
    for m in midnight_lines(xmin, xmax):
        shapes.append(dict(type="line", x0=m, x1=m, y0=0, y1=1, xref="x", yref="paper",
                           line=dict(color="rgba(0,0,0,0.35)", width=1.2, dash="dash")))

    # Axe bas (heures)
    xaxis = dict(
        title="Date / Heure",
        tickformat="%H:%M",
        dtick=dtick_ms,
        showgrid=True, gridcolor="rgba(0,0,0,0.12)",
        zeroline=False
    )

    # Axe haut (jours à minuit)
    midnights = midnight_lines(xmin, xmax)
    xaxis2 = dict(
        overlaying="x", side="top", matches="x",
        tickmode="array", tickvals=midnights, ticktext=[m.strftime("%d/%m") for m in midnights],
        showgrid=False
    )

    # petites étiquettes aux extrémités
    left_lbl  = pd.Timestamp(xmin).strftime("%d/%m %H:%M")
    right_lbl = pd.Timestamp(xmax).strftime("%d/%m %H:%M")

    fig.update_layout(
        margin=dict(l=60, r=30, t=60, b=70),
        hovermode="x",
        xaxis=xaxis, xaxis2=xaxis2,
        yaxis=dict(title=r"$L_{Aeq}$ (dB_A)", gridcolor="rgba(0,0,0,0.1)", zeroline=False),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0.01),
        shapes=shapes,
        template="plotly_white",
        title=dict(text="", x=0.5, xanchor="center")
    )
    # annotations bas-gauche / bas-droit
    fig.add_annotation(x=xmin, xref="x", y=-0.16, yref="paper", text=left_lbl,
                       showarrow=False, font=dict(size=10, color="rgba(0,0,0,0.55)"), xanchor="left")
    fig.add_annotation(x=xmax, xref="x", y=-0.16, yref="paper", text=right_lbl,
                       showarrow=False, font=dict(size=10, color="rgba(0,0,0,0.55)"), xanchor="right")

    return fig

def compute_metrics(df, removed_ts=None, diurne=(7,22), nocturne=(22,7)):
    removed_ts = set(removed_ts or [])
    t0, t1 = df.index.min(), df.index.max()
    intervals = {
        "Diurne":   build_intervals(t0,t1,diurne[0],diurne[1],["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]),
        "Nocturne": build_intervals(t0,t1,nocturne[0],nocturne[1],["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]),
    }
    out=[]
    all_concat=[]
    for name, intervs in intervals.items():
        s = concat_series(df, intervs)
        if len(s):
            s = s[~s.index.astype(str).isin(removed_ts)]
            out.append(dict(Situation=name, LAeq=round(LAeq(s),1), L90=round(LN(s,90),1), L50=round(LN(s,50),1)))
            all_concat.append(s)
    if all_concat:
        tot = pd.concat(all_concat).sort_index()
        out.append(dict(Situation="Total", LAeq=round(LAeq(tot),1), L90=round(LN(tot,90),1), L50=round(LN(tot,50),1)))
    return pd.DataFrame(out)

def fig_to_png_bytes(fig):
    return pio.to_image(fig, format="png", width=1100, height=400, scale=2)

def build_excel(fig_png_bytes, metrics_df, point_name="Point"):
    wb = Workbook(); ws = wb.active; ws.title = "Résultats"
    ws["B2"] = f"{point_name} — dBNext (Web)"; ws["B2"].font = ws["B2"].font.copy(bold=True)
    # image
    tmp = BytesIO(fig_png_bytes)
    img = XLImage(tmp); img.anchor = "B4"; ws.add_image(img)
    # metrics
    row = 25
    ws["B{}".format(row)] = "Situation"; ws["C{}".format(row)] = "LAeq (dB(A))"; ws["D{}".format(row)]="L90"; ws["E{}".format(row)]="L50"
    row += 1
    for _,r in metrics_df.iterrows():
        ws["B{}".format(row)] = r["Situation"]; ws["C{}".format(row)] = float(r["LAeq"])
        ws["D{}".format(row)] = float(r["L90"]); ws["E{}".format(row)] = float(r["L50"]); row+=1
    bio = BytesIO(); wb.save(bio); bio.seek(0); return bio.getvalue()

# =============== DASH APP ===============
app = Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])
app.title = "dBNext Web"

app.layout = dbc.Container([
    html.H3("dBNext — Analyse acoustique (Web)"),
    dbc.Row([
        dbc.Col([
            dcc.Upload(
                id='upload-data', multiple=True,
                children=html.Div(['Glissez-déposez ici ou ', html.A('choisissez des CSV')]),
                style={'width':'100%','height':'70px','lineHeight':'70px','borderWidth':'1px',
                       'borderStyle':'dashed','borderRadius':'6px','textAlign':'center','marginBottom':'8px'}
            ),
            dbc.Alert(id="upload-info", color="secondary", is_open=False, fade=True, style={"fontSize":"90%"}),
            dbc.Input(id="point-name", placeholder="Nom du point (ex. CS11-N1)", type="text", value="Point 1", className="mb-2"),
            dbc.Row([
                dbc.Col([html.Small("Diurne début (h)"), dbc.Input(id="h-diurne-deb", type="number", min=0, max=23, value=7)], width=6),
                dbc.Col([html.Small("Diurne fin (h)"),   dbc.Input(id="h-diurne-fin", type="number", min=0, max=23, value=22)], width=6),
            ], className="mb-2"),
            dbc.Row([
                dbc.Col([html.Small("Nocturne début (h)"), dbc.Input(id="h-noct-deb", type="number", min=0, max=23, value=22)], width=6),
                dbc.Col([html.Small("Nocturne fin (h)"),   dbc.Input(id="h-noct-fin", type="number", min=0, max=23, value=7)], width=6),
            ], className="mb-2"),
            dbc.InputGroup([
                dbc.InputGroupText("Seuil (dB)"),
                dbc.Input(id="threshold", type="number", step="0.1", value=999)
            ], className="mb-2"),
            dbc.ButtonGroup([
                dbc.Button("Appliquer seuil", id="btn-threshold", color="warning"),
                dbc.Button("Undo", id="btn-undo", color="secondary"),
                dbc.Button("Redo", id="btn-redo", color="secondary"),
                dbc.Button("Reset", id="btn-reset", color="secondary")
            ], className="mb-2"),
            html.Small("Astuce : utilisez l’outil *Box Select* ou *Lasso* de Plotly sur le graphique pour gommer.", className="text-muted"),
            html.Hr(),
            dbc.Button("Exporter Excel", id="btn-export", color="success"),
            dcc.Download(id="download-excel"),
        ], md=4),
        dbc.Col([
            dcc.Graph(id="graph", config={"displaylogo": False, "modeBarButtonsToAdd":["lasso2d","select2d"]}),
        ], md=8)
    ]),
    # Stores (état)
    dcc.Store(id="store-raw-df"),          # parquet JSON du df fusionné
    dcc.Store(id="store-removed-ts"),      # timestamps (str) exclus
    dcc.Store(id="store-history"),         # pile undo
    dcc.Store(id="store-future"),          # pile redo
], fluid=True, className="p-3")

# =============== CALLBACKS ===============
@app.callback(
    Output("upload-info","children"),
    Output("upload-info","is_open"),
    Output("store-raw-df","data"),
    Input("upload-data","contents"),
    State("upload-data","filename"),
    prevent_initial_call=True
)
def load_files(contents, filenames):
    if not contents: return "", False, None
    dfs=[]
    msgs=[]
    for c,f in zip(contents, filenames):
        df = lire_fichier_dBNext_from_content(c, f)
        dfs.append(df)
        msgs.append(f"✅ {f}: {len(df)} pts | {df.index.min()} → {df.index.max()}")
    df_all = pd.concat(dfs).sort_index()
    # sérialisation compacte (records + ISO)
    payload = {"t": df_all.index.astype(str).tolist(), "v": df_all["LAeq"].astype(float).round(3).tolist()}
    return html.Ul([html.Li(m) for m in msgs]), True, payload

@app.callback(
    Output("graph","figure"),
    Input("store-raw-df","data"),
    Input("store-removed-ts","data"),
    Input("h-diurne-deb","value"), Input("h-diurne-fin","value"),
    Input("h-noct-deb","value"),   Input("h-noct-fin","value"),
    prevent_initial_call=True
)
def update_graph(payload, removed_ts, hd, hf, nd, nf):
    if not payload: return go.Figure()
    df = pd.DataFrame({"time": pd.to_datetime(payload["t"]), "LAeq": payload["v"]}).set_index("time")
    fig = build_figure(df, removed_ts, diurne=(hd,hf), nocturne=(nd,nf))
    return fig

# Gomme par sélection (box/lasso)
@app.callback(
    Output("store-removed-ts","data"),
    Output("store-history","data"),
    Output("store-future","data"),
    Input("graph","selectedData"),
    State("store-removed-ts","data"),
    State("store-history","data"),
    State("store-future","data"),
    prevent_initial_call=True
)
def erase_selected(selectedData, removed_ts, history, future):
    if not selectedData: return removed_ts, history, future
    removed_ts = set(removed_ts or [])
    history = (history or []) + [sorted(removed_ts)]  # push
    future = []  # clear redo
    # selected points → timestamps
    for p in selectedData.get("points", []):
        # p["x"] est déjà une date ISO
        removed_ts.add(str(p["x"]))
    return sorted(removed_ts), history, future

# Seuil / Undo / Redo / Reset
@app.callback(
    Output("store-removed-ts","data"),
    Output("store-history","data"),
    Output("store-future","data"),
    Input("btn-threshold","n_clicks"),
    Input("btn-undo","n_clicks"),
    Input("btn-redo","n_clicks"),
    Input("btn-reset","n_clicks"),
    State("threshold","value"),
    State("store-raw-df","data"),
    State("store-removed-ts","data"),
    State("store-history","data"),
    State("store-future","data"),
    prevent_initial_call=True
)
def actions(th, undo, redo, reset, thr_value, payload, removed_ts, history, future):
    trig = callback_context.triggered[0]["prop_id"].split(".")[0]
    removed_ts = set(removed_ts or [])
    history = history or []; future = future or []
    if trig == "btn-threshold":
        if not payload: return sorted(removed_ts), history, future
        # push history
        history = history + [sorted(removed_ts)]
        future = []
        df = pd.DataFrame({"time": pd.to_datetime(payload["t"]), "LAeq": payload["v"]}).set_index("time")
        m = df["LAeq"] >= float(thr_value)
        removed_ts |= set(df.index[m].astype(str).tolist())
    elif trig == "btn-undo" and history:
        future = future + [sorted(removed_ts)]
        removed_ts = set(history.pop())
    elif trig == "btn-redo" and future:
        history = history + [sorted(removed_ts)]
        removed_ts = set(future.pop())
    elif trig == "btn-reset":
        if removed_ts:
            history = history + [sorted(removed_ts)]
        removed_ts = set()
        future = []
    return sorted(removed_ts), history, future

# Export Excel
@app.callback(
    Output("download-excel","data"),
    Input("btn-export","n_clicks"),
    State("store-raw-df","data"),
    State("store-removed-ts","data"),
    State("h-diurne-deb","value"), State("h-diurne-fin","value"),
    State("h-noct-deb","value"),   State("h-noct-fin","value"),
    State("point-name","value"),
    prevent_initial_call=True
)
def export_excel(n, payload, removed_ts, hd, hf, nd, nf, point_name):
    if not payload: return None
    df = pd.DataFrame({"time": pd.to_datetime(payload["t"]), "LAeq": payload["v"]}).set_index("time")
    fig = build_figure(df, removed_ts, diurne=(hd,hf), nocturne=(nd,nf))
    png_bytes = fig_to_png_bytes(fig)
    metrics = compute_metrics(df, removed_ts, diurne=(hd,hf), nocturne=(nd,nf))
    xlsx_bytes = build_excel(png_bytes, metrics, point_name=point_name or "Point")
    filename = f"Analyse_Acoustique_{(point_name or 'Point').replace(' ','_')}.xlsx"
    return dcc.send_bytes(lambda b: b.write(xlsx_bytes), filename)

if __name__ == "__main__":
    app.run_server(debug=True)
