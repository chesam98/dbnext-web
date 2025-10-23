\
# dBNext — coeur métier pour l'app web (Dash)
# Portage du script bureau (v7.2) vers un module serveur web
import io, os, tempfile
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")  # rendu headless côté serveur
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import MultipleLocator
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage

# ---------- PALETTES / STYLES ----------
PALETTE = {"Diurne": "#1f77b4", "Nocturne": "#ff7f0e", "Maintenance": "#2ca02c"}
HEADER_FILL = {"Diurne":"B7D5F2", "Nocturne":"F7C38C", "Maintenance":"A4D8A4"}
FILL_RED = "E74C3C"    # 30 min bruyantes
FILL_BLUE = "5DADE2"   # 30 min calmes
FILL_TOTAL = "E2F0D9"
GRID_GRAY = "808080"
JOUR_IDX = {"lundi":0,"mardi":1,"mercredi":2,"jeudi":3,"vendredi":4,"samedi":5,"dimanche":6}

# ============================================================
# 2) 🗂️ LECTURE FICHIERS (générique dBNext)
# ============================================================
def lire_fichier_dBNext_from_text(text):
    """
    Lecture générique de fichiers dBNext/CESVA déjà nettoyés depuis texte CSV.
      - 1ère colonne = temps (date/heure)
      - 2e colonne   = LAeq (float)
    """
    import io
    df = pd.read_csv(io.StringIO(text), sep=None, engine="python", encoding="utf-8")
    if df.shape[1] < 2:
        raise ValueError("❌ Le fichier doit contenir au moins deux colonnes (temps + LAeq).")

    df = df.iloc[:, :2].copy()
    df.columns = ["time_block", "LAeq"]
    df["time_block"] = pd.to_datetime(df["time_block"], errors="coerce", dayfirst=True)
    df["LAeq"] = pd.to_numeric(df["LAeq"], errors="coerce")
    df = df.dropna(subset=["time_block", "LAeq"]).set_index("time_block").sort_index()
    return df

# ============================================================
# 3) 🧭 UTILITAIRES TEMPS & SÉRIES
# ============================================================
def normalize_jours(jours):
    return [JOUR_IDX[j.lower()] for j in (jours or []) if j.lower() in JOUR_IDX]

def build_intervals(t0, t1, h_deb, h_fin, jours=None):
    days = normalize_jours(jours)
    cur = datetime(t0.year, t0.month, t0.day)
    out=[]
    while cur < t1:
        if (not days) or (cur.weekday() in days):
            if h_deb <= h_fin:
                s = cur.replace(hour=h_deb, minute=0); e = cur.replace(hour=h_fin, minute=0)
            else:
                s = cur.replace(hour=h_deb, minute=0)
                e = (cur+timedelta(days=1)).replace(hour=h_fin, minute=0)
            if s < t1 and e > t0: out.append((max(s,t0), min(e,t1)))
        cur += timedelta(days=1)
    return out

def concat_series(df, intervals, col="LAeq"):
    segs = [df.loc[(df.index>=s)&(df.index<=e), col] for s,e in intervals]
    return pd.concat(segs).sort_index() if segs else pd.Series(dtype=float)

def concat_series_with_gaps(df, intervals, col="LAeq"):
    chunks=[]
    for s,e in intervals:
        sub=df.loc[(df.index>=s)&(df.index<=e), col]
        if not sub.empty:
            chunks.append(sub)
            chunks.append(pd.Series([np.nan], index=[sub.index[-1]+pd.Timedelta(seconds=1)]))
    return pd.concat(chunks).sort_index() if chunks else pd.Series(dtype=float)

# ============================================================
# 4) 🔊 ACOUSTIQUE
# ============================================================
def LAeq(series):
    return float(10*np.log10(np.nanmean(10**(series/10)))) if len(series)>0 else float("nan")

def LN(series,N):
    return float(np.nanpercentile(series,100-N)) if len(series)>0 else float("nan")

def infer_dt_seconds(series):
    if len(series.index)<2: return 1
    dt=pd.Series(series.index).diff().dt.total_seconds().median()
    return int(max(1, round(dt)))

def rolling_LAeq(series, window_s=1800):
    if series.empty: return pd.Series(dtype=float)
    dt=infer_dt_seconds(series); n=max(1,int(window_s//dt))
    vals=10**(series.values/10)
    if len(vals)<n: return pd.Series(dtype=float)
    ener=np.convolve(vals, np.ones(n), "valid")/n
    idx=series.index[n-1:]
    return pd.Series(10*np.log10(ener), index=idx)

def rolling_extrema(series, window_s=1800, mode="max"):
    roll=rolling_LAeq(series, window_s)
    if roll.empty: return None,None,float("nan")
    idx = roll.idxmax() if mode=="max" else roll.idxmin()
    return idx - pd.Timedelta(seconds=window_s), idx, float(roll.loc[idx])

def total_duration(intervals):
    return str(sum((e-s for s,e in intervals), timedelta()))

# ============================================================
# 6) 📊 EXPORT PNG (lisible) — pour insertion Excel
# ============================================================
def _configure_time_axis_export(ax, xmin, xmax):
    import pandas as pd
    import matplotlib.dates as mdates

    span = pd.Timestamp(xmax) - pd.Timestamp(xmin)
    if span <= pd.Timedelta("6H"):
        hr_int, fmt_bottom = 1, "%H:%M"
    elif span <= pd.Timedelta("24H"):
        hr_int, fmt_bottom = 2, "%H:%M"
    elif span <= pd.Timedelta("3D"):
        hr_int, fmt_bottom = 3, "%H:%M"
    else:
        hr_int, fmt_bottom = 6, "%Hh"

    ax.xaxis.set_major_locator(mdates.HourLocator(interval=hr_int))
    ax.xaxis.set_major_formatter(mdates.DateFormatter(fmt_bottom))
    ax.xaxis.set_minor_locator(mdates.HourLocator(interval=1))

    ax_top = ax.twiny()
    ax_top.set_xlim(ax.get_xlim())
    day_loc = mdates.DayLocator()  # tick garanti à 00:00
    ax_top.xaxis.set_major_locator(day_loc)
    ax_top.xaxis.set_major_formatter(mdates.DateFormatter("%d/%m"))
    ax_top.tick_params(axis="x", top=True, labeltop=True, bottom=False, labelbottom=False, labelsize=8, pad=4)

    ax.grid(True, which="major", ls="--", alpha=0.55, zorder=1)
    ax.grid(True, which="minor", ls=":",  alpha=0.25, zorder=1)

    start_mid = pd.Timestamp(xmin).floor("D")
    end_mid   = pd.Timestamp(xmax).ceil("D")
    for j in pd.date_range(start_mid, end_mid, freq="D"):
        ax.axvline(j, color="black", linestyle="--", alpha=0.35, lw=1.1, zorder=6)

    fmt = "%d/%m %H:%M"
    left_lbl  = pd.Timestamp(xmin).strftime(fmt)
    right_lbl = pd.Timestamp(xmax).strftime(fmt)
    for x, txt, ha in [(0.0, left_lbl, "left"), (1.0, right_lbl, "right")]:
        ax.text(x, -0.12, txt, transform=ax.transAxes,
                ha=ha, va="top", fontsize=7, color="0.35",
                alpha=0.95, clip_on=False)
    return ax_top

def make_plot_png_bytes(series_map, title):
    """
    Fabrique un PNG en mémoire depuis un dict label→(x, y, color).
    Retourne bytes PNG.
    """
    import pandas as pd
    import numpy as np

    plt.figure(figsize=(8.8, 3.2), dpi=150, constrained_layout=True)
    ax = plt.gca()
    all_x = []
    for label, (x, y, color) in series_map.items():
        if len(x):
            if label == "Perturbations":
                ax.scatter(x, y, s=16, color="red", label="Perturbations", zorder=5)
            else:
                ax.plot(x, y, lw=1.2, color=color, label=label, zorder=3)
            all_x.extend(pd.to_datetime(x))

    if not all_x:
        ax.text(0.5, 0.5, "Aucune donnée", ha="center", va="center", transform=ax.transAxes)
        bio = io.BytesIO()
        plt.savefig(bio, format="png", bbox_inches="tight")
        plt.close()
        return bio.getvalue()

    xmin, xmax = min(all_x), max(all_x)
    ax.set_xlim(xmin, xmax)

    ax.yaxis.set_major_locator(MultipleLocator(5))
    ax.yaxis.set_minor_locator(MultipleLocator(1))

    _configure_time_axis_export(ax, xmin, xmax)
    ax.set_title(title, fontsize=11)
    ax.set_ylabel(r"$L_{Aeq}$ (dB$_A$)")
    ax.set_xlabel("Date / Heure")
    ax.legend(frameon=False, fontsize=8, loc="upper left", ncol=1)

    bio = io.BytesIO()
    plt.savefig(bio, format="png", bbox_inches="tight")
    plt.close()
    return bio.getvalue()

# ============================================================
# 7) 📄 FEUILLE PAR POINT — export Excel
# ============================================================
def style_cell(cell, bold=False, color=None, align="center"):
    cell.font = Font(name="Century Gothic", size=11, bold=bold, color=("FFFFFF" if color else "000000"))
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if color: cell.fill = PatternFill("solid", fgColor=color)

def add_table(ws, start_row, start_col, title, df, header_color=None):
    ncols=len(df.columns)
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+ncols-1)
    tcell=ws.cell(row=start_row, column=start_col, value=title)
    style_cell(tcell, bold=True, align="center")
    thin=Side(style="medium", color=GRID_GRAY)

    for j,h in enumerate(df.columns):
        c=ws.cell(row=start_row+1, column=start_col+j, value=h)
        style_cell(c, bold=True, color=header_color); c.border=Border(top=thin,left=thin,right=thin,bottom=thin)
    for i in range(len(df)):
        for j,h in enumerate(df.columns):
            v=df.iloc[i,j]; c=ws.cell(row=start_row+2+i, column=start_col+j, value=v)
            try: float(v); style_cell(c, align="right")
            except: style_cell(c, align="center")
            c.border=Border(top=thin,left=thin,right=thin,bottom=thin)
    return start_row+2+len(df)

def process_point_to_sheets(wb, point_name, point_type, df_global, situations, cleaned_masks):
    """
    Construit les feuilles pour un point :
      - Graphiques Jour/Nuit + autres situations (PNGs insérés)
      - Tableaux LAeq/L90/L50 sur séries nettoyées
      - 30 min caractéristiques (selon type : LP -> max, ZER -> min)
      - Période totale
    cleaned_masks: dict {situation_name: index list to exclude}
    """
    types = ["LP","ZER"] if point_type=="HYBRIDE" else [point_type]

    # Séries par situation depuis le BRUT
    series_sit = {nom: concat_series(df_global, intervs, col="LAeq") for nom,intervs in situations.items()}

    # Appliquer masques nettoyés
    cleaned_perturb = {}  # pour scatter rouge
    for nom, s in series_sit.items():
        if s.empty: continue
        mask_idx = cleaned_masks.get(nom, [])
        if mask_idx:
            s_clean = s.copy()
            s_clean.loc[mask_idx] = np.nan  # exclus
            series_sit[nom] = s_clean
            cleaned_perturb[nom] = s.loc[mask_idx].dropna()

    for t in types:
        ws = wb.create_sheet(f"{t} - {point_name}")

        # Graphiques
        dn_series={}; extra_series={}
        perturb_dn_x=[]; perturb_dn_y=[]
        for k in ["Diurne","Nocturne"]:
            if k in situations:
                y_gap = concat_series_with_gaps(df_global, situations[k], col="LAeq")
                if not y_gap.empty:
                    dn_series[k] = (y_gap.index, y_gap.values, PALETTE.get(k, "#000000"))
                if k in cleaned_perturb and not cleaned_perturb[k].empty:
                    perturb_dn_x.extend(cleaned_perturb[k].index); perturb_dn_y.extend(cleaned_perturb[k].values)

        if perturb_dn_x:
            dn_series["Perturbations"] = (pd.Index(perturb_dn_x), np.array(perturb_dn_y), "red")

        perturb_ex_x=[]; perturb_ex_y=[]
        for k, intervs in situations.items():
            if k in ["Diurne","Nocturne"]: continue
            y_gap = concat_series_with_gaps(df_global, intervs, col="LAeq")
            if not y_gap.empty:
                extra_series[k] = (y_gap.index, y_gap.values, PALETTE.get(k,"#333333"))
            if k in cleaned_perturb and not cleaned_perturb[k].empty:
                perturb_ex_x.extend(cleaned_perturb[k].index); perturb_ex_y.extend(cleaned_perturb[k].values)
        if perturb_ex_x:
            extra_series["Perturbations"] = (pd.Index(perturb_ex_x), np.array(perturb_ex_y), "red")

        dn_png = make_plot_png_bytes(dn_series or {"Global":(df_global.index, df_global["LAeq"].values, "#1f77b4")}, f"{point_name} | Type : {t}")
        ex_png = make_plot_png_bytes(extra_series or {" ": (pd.Index([]),[], "#2ca02c")}, f"Autres situations — {point_name}")

        # Insérer images
        import tempfile
        tmp1 = tempfile.NamedTemporaryFile(delete=False, suffix=".png"); tmp1.write(dn_png); tmp1.flush()
        tmp2 = tempfile.NamedTemporaryFile(delete=False, suffix=".png"); tmp2.write(ex_png); tmp2.flush()

        ws.add_image(XLImage(tmp1.name), "B2")
        ws.add_image(XLImage(tmp2.name), "J2")

        # Tableaux (calculs sur séries_sit NETTOYÉES)
        row0=24; col0=2
        per_sit_main={}
        for nom, intervs in situations.items():
            s = series_sit.get(nom, pd.Series(dtype=float))
            rows=[]
            if not s.empty:
                rows.append({"Situation":nom,"LAeq (dB(A))":round(LAeq(s),1),
                             "L90":round(LN(s,90),1),"L50":round(LN(s,50),1),
                             "Durée":total_duration(intervs)})
            per_sit_main[nom] = pd.DataFrame(rows, columns=["Situation","LAeq (dB(A))","L90","L50","Durée"])

        col=col0; max_end=row0
        ordered = ["Diurne","Nocturne"] + [k for k in situations.keys() if k not in ["Diurne","Nocturne"]]
        for nom in ordered:
            df = per_sit_main.get(nom)
            if df is not None and not df.empty:
                end = add_table(ws, row0, col, f"Situation : {nom}", df, HEADER_FILL.get(nom))
                max_end=max(max_end,end)
                col += len(df.columns)+1

        # Période totale
        if series_sit:
            seg_total = pd.concat(series_sit.values())
            seg_total = seg_total.groupby(seg_total.index).first().sort_index()
        else:
            seg_total = pd.Series(dtype=float)

        tot = pd.DataFrame([{
            "LAeq (dB(A))": round(LAeq(seg_total),1) if not seg_total.empty else np.nan,
            "L90": round(LN(seg_total,90),1) if not seg_total.empty else np.nan,
            "L50": round(LN(seg_total,50),1) if not seg_total.empty else np.nan,
            "Durée": total_duration(sum(situations.values(), []))
        }])
        add_table(ws, max_end+4, col0, "Période totale", tot, FILL_TOTAL)

        # 30 minutes caractéristiques
        rows=[]
        for nom, s in series_sit.items():
            if s.empty: continue
            if t=="LP":
                st,en,_ = rolling_extrema(s, 1800, "max"); label="30 min plus bruyantes"; color=FILL_RED
            else:
                st,en,_ = rolling_extrema(s, 1800, "min"); label="30 min plus calmes"; color=FILL_BLUE
            if st:
                seg = s.loc[(s.index>=st)&(s.index<=en)]
                rows.append({"Situation":nom,"Début":st,"Fin":en,"Durée":"00:30:00",
                             "LAeq (dB(A))":round(LAeq(seg),1),"L90":round(LN(seg,90),1),"L50":round(LN(seg,50),1)})
        if rows:
            ext = pd.DataFrame(rows)
            add_table(ws, max_end+4, col0+12, label, ext, color)

        # Titre
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=22)
        tcell = ws.cell(row=1, column=2, value=f"{point_name} — Type : {t}   (dBNext)")
        style_cell(tcell, bold=True, align="center")

def build_workbook(points):
    """
    points: liste de dicts:
      {
        "name": str,
        "type": "LP"|"ZER"|"HYBRIDE",
        "df": DataFrame(index=datetime, LAeq col),
        "situations": dict[str, list[(start,end)]],
        "cleaned_masks": dict[situation_name -> list(datetime index to exclude)]
      }
    Retourne bytes XLSX.
    """
    wb = Workbook()
    wb.active.title = "Points de mesure"
    wb.create_sheet("Annexe")

    for p in points:
        process_point_to_sheets(wb, p["name"], p["type"], p["df"], p["situations"], p.get("cleaned_masks", {}))

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
